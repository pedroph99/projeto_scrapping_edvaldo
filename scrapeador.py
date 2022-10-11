import os
import time
import selenium
import openpyxl
import undetected_chromedriver as uc

from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from openpyxl.utils.cell import get_column_letter
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#chrome_options.add_argument('--headless')
chrome_options = webdriver.ChromeOptions()


def scrapea(cnpj):
    delay = 10
    driver = uc.Chrome()
    # driver2 = driver.get('https://consopt.www8.receita.fazenda.gov.br/consultaoptantes')

    time.sleep(3)
    input_text = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.CLASS_NAME, 'form-control')))

    time.sleep(3)
    input_text = WebDriverWait(driver, delay).until(EC.presence_of_element_located((By.CLASS_NAME, 'form-control')))
    input_text.send_keys(cnpj)
    button_text = driver.find_element(by=By.CLASS_NAME, value='btn-verde')

    time.sleep(1)
    button_text.click()
    time.sleep(2)
    WebDriverWait(driver, delay * 4).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'spanValorVerde')))
    optante = driver.find_elements(by=By.CLASS_NAME, value='spanValorVerde')

    return optante[2].text


def ler_excell(file, cnpj):
    # Lê planilhas e identifica a última coluna
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    target_column = folha.max_column

    # escreve o simples nacional caso não haja o nome ainda na planilha....
    escreve_simples_nacional(file)
    # identifica se planilha já tem informação do simples ou não
    if folha['%s1' % (get_column_letter(target_column))].value != 'Simples nacional':
        target_column = folha.max_column + 1

    # pega a letra do target_column
    # target_letter = get_column_letter(target_column)

    # Algoritmo que detecta valores em branco na planilha, identificando, portanto, o último elemento
    counter = 1
    last_value = None
    testa_ultimo_valor = True
    while testa_ultimo_valor:
        if folha['%s%s' % (get_column_letter(target_column), counter)].value:
            counter += 1
        else:
            last_value = '%s%s' % (get_column_letter(target_column), counter)
            testa_ultimo_valor = False

    # fecha a planilha e começa o processo de modificação do excell
    planilha.close()

    escreve_excell(file, last_value, cnpj)


def escreve_excell(file, last_value, cnpj):
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    print(last_value)
    folha[last_value] = scrapea(cnpj)
    planilha.save('empresas.xlsx')


def testa_cnpj(cnpj):
    teste = cnpj
    teste2 = teste.replace('.', '').replace('/', '').replace('-', '')
    return teste2


def escreve_simples_nacional(file):
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    target_column = folha.max_column
    if folha['%s1' % (get_column_letter(target_column))].value != 'Simples nacional':
        folha['%s1' % (get_column_letter(target_column + 1))] = 'Simples nacional'
        planilha.save()


def pega_cnpjs(file):
    wb = openpyxl.load_workbook(file)
    listacnpj = wb.sheetnames
    ws = wb[listacnpj[0]]

    # Identifica o campo CNPJ
    trace_row = None
    trace_column = None

    for x in range(ws.max_column + 1):
        for y in range(ws.max_row + 1):
            try:
                d = ws.cell(row=y, column=x)
                if d.value.lower() == 'cnpj':
                    trace_row = y
                    trace_column = x
            except:
                pass
    # pega todos os CNPJS de maneira vertical.
    lista_cnpj = []

    current_row = int(trace_row) + 1
    while current_row <= int(ws.max_row):
        d = ws.cell(row=current_row, column=trace_column)
        if d.value is not None:
            lista_cnpj.append(d.value)
        else:
            pass

        current_row += 1
    
    return lista_cnpj


def pega_ultimo_preenchido(file):
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    target_column = folha.max_column
    counter = 1
    testa_ultimo_valor = True
    last_value = None
    while testa_ultimo_valor:
        if folha['%s%s' % (get_column_letter(target_column), counter)].value:
            counter += 1
        else:
            last_value = '%s%s' % (get_column_letter(target_column), counter)
            testa_ultimo_valor = False

    planilha.close()
    original_list = pega_cnpjs(file)
    del original_list[0:(counter - 2)]

    return original_list


def preenche_erro(file):
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    target_column = folha.max_column
    counter = 1
    testa_ultimo_valor = True
    last_value = None
    while testa_ultimo_valor:
        if folha['%s%s' % (get_column_letter(target_column), counter)].value:
            counter += 1
        else:
            last_value = '%s%s' % (get_column_letter(target_column), counter)
            testa_ultimo_valor = False
    print(last_value)
    folha[last_value] = 'Erro. Não foi possível pegar dados desta empresa.'
    planilha.save('empresas.xlsx')
    planilha.close()


def lista_erros(file):
    wb= openpyxl.load_workbook( file)

    listacnpj=wb.sheetnames

    ws=wb[listacnpj[0]]
    #Identifica o campo simples nacional
    trace_row=None
    trace_column=None

    for x in range(ws.max_column+1):
        for y in range(ws.max_row+1):
            try:
                d=ws.cell(row=y, column=x)
                try:
                    
                    if d.value.lower() == 'simples nacional':
                        trace_row=y
                        trace_column=x
                except:
                    break

            except:
                pass
       #pega todos os CNPJS de maneira vertical. 
    lista_erros=[]

    current_row=int(trace_row) +1
    while current_row <= int(ws.max_row):
        d=ws.cell(row=current_row, column=trace_column)
        try:
            if d.value.lower().startswith('erro'):
                lista_erros.append((current_row, trace_column ))
            else:
                pass
        except:
            pass

        current_row+=1

    print(lista_erros)
    
    return lista_erros


def pega_cnpjs_com_erros(file):
    wb= openpyxl.load_workbook( file)

    listacnpj=wb.sheetnames

    ws=wb[listacnpj[0]]
    #Identifica o campo CNPJ
    trace_row=None
    trace_column=None
    for x in range(ws.max_column+1):
        for y in range(ws.max_row+1):
            try:
                d=ws.cell(row=y, column=x)
                if d.value.lower() == 'cnpj':
                    trace_column=x
                
                    return x
            except:
                pass

def corrige_erros(file):
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    print(pega_cnpjs_com_erros(file))
    coordenadas = lista_erros(file)
    for x in coordenadas:
        print(x)
        cnpj_erro = folha.cell(row=x[0], column=x[1]-1).value
        print(cnpj_erro)
        print(pega_cnpjs_com_erros(file))
        planilha.close()
        escreve_excell(file, '%s%s'%(get_column_letter(x[1]), x[0]), cnpj_erro)




    


if __name__ == '__main__':
   # for x in pega_cnpjs('empresas.xlsx'):
   #     ler_excell('empresas.xlsx', testa_cnpj(x))
   
   #for x in pega_ultimo_preenchido('empresas.xlsx'):
   # time.sleep(2)
   # try:
   #     ler_excell('empresas.xlsx', testa_cnpj(x))
   # except:
   #     preenche_erro('empresas.xlsx')
   lista_erros('empresas.xlsx')
   corrige_erros('empresas.xlsx')


    

    


