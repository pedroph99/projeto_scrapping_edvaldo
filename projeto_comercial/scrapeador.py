import os
from termios import OPOST
import time
import selenium
import openpyxl
import undetected_chromedriver as uc

from openpyxl.utils.cell import get_column_letter
from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def pega_ultimo_preenchido(file):
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    target_column = folha.max_column
    counter = 1
    testa_ultimo_valor = True
    last_value = None
    while testa_ultimo_valor:
        if folha['%s%s' % (get_column_letter(target_column), counter)].value:

            counter += 1

        else:

            last_value = '%s%s' % (get_column_letter(target_column), counter)
            testa_ultimo_valor = False

    planilha.close()
    original_list = pega_cnpjs(file)
    del original_list[0:(counter-2)]

    return original_list


def pega_cnpjs(file):
    wb = openpyxl.load_workbook(file)
    listacnpj = wb.sheetnames
    ws = wb[listacnpj[0]]
    # Identifica o campo CNPJ
    trace_row = None
    trace_column = None
    for x in range(ws.max_column+1):
        for y in range(ws.max_row+1):
            try:
                d = ws.cell(row=y, column=x)
                if d.value.lower() == 'cnpj':
                    trace_row = y
                    trace_column = x
            except:
                pass
       # pega todos os CNPJS de maneira vertical.
    lista_cnpj = []
    current_row = int(trace_row) + 1
    while current_row <= int(ws.max_row):
        d = ws.cell(row=current_row, column=trace_column)
        if d.value != None:

            lista_cnpj.append(d.value)
        else:
            pass

        current_row += 1

    return lista_cnpj


def ler_excell(file, cnpj):
    # Lê planilhas e identifica a última coluna
    escreve_ISS(file)
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    target_column = folha.max_column
    # identifica se planilha já tem informação do simples ou não
    if folha['%s1' % (get_column_letter(target_column))].value != 'ISS':
        target_column = folha.max_column + 1
    # pega a letra do target_column
    target_letter = get_column_letter(target_column)

    # Algoritmo que detecta valores em branco na planilha, identificando, portanto, o último elemento
    last_value = None
    testa_ultimo_valor = True
    counter = 1
    while testa_ultimo_valor:
        if folha['%s%s' % (get_column_letter(target_column), counter)].value:
            counter += 1
        else:
            last_value = '%s%s' % (get_column_letter(target_column), counter)
            testa_ultimo_valor = False
    # fecha a planilha e começa o processo de modificação do excell
    planilha.close()
    escreve_excell(file, last_value, cnpj)


def escreve_excell(file, last_value, cnpj):
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    print(last_value)
    folha[last_value] = scrapea(cnpj)
    planilha.save(file)


def escreve_ISS(file):
    planilha = openpyxl.load_workbook(file)
    folha = planilha.active
    target_column = folha.max_column
    print(target_column)
    if folha['%s1' % (get_column_letter(target_column))].value != 'ISS':
        print('ok')
        print('%s1' % (get_column_letter(target_column+1)))
        print('ok')
        folha['%s1' % (get_column_letter(target_column+1))] = 'ISS'
        planilha.save(filename=file)
    planilha.close()


def scrapea(cnpj):
    opcoes = uc.ChromeOptions()
    opcoes.add_argument("--headless")
    driver = uc.Chrome(options=opcoes)
    delay = 10
    driver2 = driver.get(
        'https://redesim.jucepe.pe.gov.br/regin.externo/CON_ViabilidadeSelecaoExternoV4.aspx?')
    forms = driver.find_element(By.ID, 'txtCNPJBusca')
    forms.send_keys(cnpj)
    botao = driver.find_element(By.ID, 'btnBuscar')
    botao.click()

    time.sleep(3)
    try:

        links_dados = driver.find_elements(By.CLASS_NAME, 'dados')
        
        if len(driver.find_elements(By.CLASS_NAME, 'AlternativeDataList')) > 0:
            for y in driver.find_elements(By.CLASS_NAME, 'AlternativeDataList'):
                links_dados.append(y)
        for x in links_dados:
            tds = x.find_elements(By.TAG_NAME, 'td')
            link_final = tds[-1].find_elements(By.TAG_NAME, 'a')[-1]
            link_final.click()
            time.sleep(10)
            
            try:
                time.sleep(4)
                print('Tentando pegar o texto')
                driver.switch_to.window(driver.window_handles[1])
                texto_objetivo = driver.find_element(By.ID, 'txtISS')
                print(texto_objetivo.text)
                return texto_objetivo.text
                time.sleep(5)
            except:
                driver.switch_to.window(driver.window_handles[1])
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(3)
        try:
            texto_objetivo = driver.find_element(By.ID, 'txtISS')
            print(texto_objetivo.text)
            return texto_objetivo.text
        except:
            print('Empresa não possui ISS')
            return 'Empresa não possui ISS'
    except:
        try:
            texto_objetivo = driver.find_element(By.ID, 'txtISS')
            print(texto_objetivo.text)
            return texto_objetivo.text
        except:
            print('Erro.. Empresa sem ISS')
            return 'Erro.. Empresa sem ISS E/OU SEM REGISTRO!'
    driver.quit()


if __name__ == '__main__':
    for x in pega_ultimo_preenchido('empresas.xlsx'):
        print(x)
        ler_excell('empresas.xlsx', x)
        time.sleep(3)
